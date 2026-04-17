"""
STEP 3.6 — LLM relevance gate (Gemini 2.5 Flash, free tier).

For each candidate vendor, fetch the homepage text and ask Gemini:
  1. Does this company serve the self-storage industry?  (YES / NO / UNCLEAR)
  2. Which category from our taxonomy fits best?
  3. One-sentence reason.

We also apply a fast pre-filter (before any API calls) to drop .gov / .mil / .edu
domains and anything already in the Supabase vendors table.

Resumable: re-running the script skips domains already written to the output
files. Designed for the free tier (~10 requests/minute, ~250-500/day for
gemini-2.5-flash): if you can't finish in one session, run again tomorrow and
it picks up where it left off.

Input:    templates/03_5_vendor_candidates.xlsx
Outputs:  templates/03_6_relevant.xlsx       (YES)
          templates/03_6_not_relevant.xlsx   (NO)
          templates/03_6_needs_review.xlsx   (UNCLEAR, fetch error, or API error)

CLI:
    python scripts/leads_cleanup/03_6_relevance_gate.py                # full run
    python scripts/leads_cleanup/03_6_relevance_gate.py --limit 20     # smoke test
"""

from __future__ import annotations

import argparse
import asyncio
import json
import os
import random
import sys
import time
from pathlib import Path
from typing import Any

import httpx
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
BASE_DIR = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates"
)
INPUT = BASE_DIR / "03_5_vendor_candidates.xlsx"
OUT_RELEVANT = BASE_DIR / "03_6_relevant.xlsx"
OUT_NOT_RELEVANT = BASE_DIR / "03_6_not_relevant.xlsx"
OUT_REVIEW = BASE_DIR / "03_6_needs_review.xlsx"

ENV_PATH = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor\.env.local"
)

GEMINI_MODEL = "gemini-2.5-flash"
GEMINI_ENDPOINT = (
    f"https://generativelanguage.googleapis.com/v1beta/models/{GEMINI_MODEL}:generateContent"
)

# Paid tier 1 for gemini-2.5-flash: 1,000 RPM / 10,000 RPD.
# Natural fetch+API time is ~4-7s per row, so a 1s floor is just a safety net
# that almost never fires in practice.
REQUEST_INTERVAL_SECONDS = 1.0
FETCH_TIMEOUT = 15
GEMINI_TIMEOUT = 60
MAX_PAGE_TEXT_CHARS = 3500   # keep prompt small; first ~3.5k chars is plenty

# Same TLD rules as step 2 (in case anything slipped through an earlier run)
BLOCKED_TLD_SUFFIXES = (
    ".gov", ".mil", ".edu",
    ".gov.uk", ".gov.au", ".gov.ca", ".gov.nz", ".gov.in", ".gov.za",
    ".gov.sg", ".gov.ie", ".gov.hk", ".gov.jp", ".gov.tw", ".gov.il",
    ".gov.br", ".gov.mx", ".gouv.fr",
    ".gc.ca",
    ".mil.uk", ".mil.au", ".mil.nz", ".mil.za", ".mil.br",
    ".edu.au", ".edu.cn", ".edu.hk", ".edu.sg", ".edu.in", ".edu.tw",
    ".edu.mx", ".edu.br", ".edu.my",
    ".ac.uk", ".ac.nz", ".ac.in", ".ac.jp", ".ac.za", ".ac.kr", ".ac.ir", ".ac.il",
)


USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]


# ---------- env loader ----------
def load_env_local():
    if not ENV_PATH.exists():
        return
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


# ---------- Supabase helpers ----------
def fetch_categories() -> list[str]:
    """Load category_slug list from Supabase."""
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        raise RuntimeError("Supabase env vars missing; cannot fetch category list.")
    with httpx.Client(timeout=30) as c:
        r = c.get(
            f"{url}/rest/v1/categories",
            params={"select": "slug", "order": "slug.asc"},
            headers={"apikey": key, "Authorization": f"Bearer {key}"},
        )
        r.raise_for_status()
        rows = r.json()
    return [row["slug"] for row in rows if row.get("slug")]


def fetch_existing_vendor_domains() -> set[str]:
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        return set()
    from urllib.parse import urlparse
    out: set[str] = set()
    offset = 0
    page = 1000
    with httpx.Client(timeout=30) as c:
        while True:
            r = c.get(
                f"{url}/rest/v1/vendors",
                params={"select": "website", "limit": str(page), "offset": str(offset)},
                headers={"apikey": key, "Authorization": f"Bearer {key}"},
            )
            r.raise_for_status()
            rows = r.json()
            if not rows:
                break
            for row in rows:
                w = row.get("website") or ""
                if not w:
                    continue
                try:
                    h = urlparse(w).netloc.lower().removeprefix("www.")
                except Exception:
                    continue
                if h:
                    out.add(h)
            if len(rows) < page:
                break
            offset += page
    return out


# ---------- xlsx helpers ----------
def read_xlsx(path: Path) -> tuple[list[str], list[list]]:
    if not path.exists():
        return [], []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    first = next(ws.iter_rows(min_row=1, max_row=1), None)
    if first is None:
        return [], []
    header = [c.value for c in first]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
    return header, rows


def write_xlsx(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(str(h)) + 4, 40))
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    wb.save(path)


# ---------- page fetch ----------
def _browser_headers(domain: str) -> dict:
    ua = random.choice(USER_AGENTS)
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Ch-Ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Referer": f"https://www.google.com/search?q={domain}",
    }


def fetch_page_text(client: httpx.Client, domain: str, url: str) -> tuple[str, str]:
    """Return (page_text, error). Page text is stripped, lower-trimmed, clipped."""
    target = url or f"https://{domain}"
    try:
        r = client.get(target, follow_redirects=True, headers=_browser_headers(domain), timeout=FETCH_TIMEOUT)
        if r.status_code >= 400:
            return "", f"http {r.status_code}"
        soup = BeautifulSoup(r.text, "html.parser")
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text(" ", strip=True)
        # Collapse whitespace
        text = " ".join(text.split())
        return text[:MAX_PAGE_TEXT_CHARS], ""
    except Exception as e:
        return "", f"{type(e).__name__}: {str(e)[:120]}"


# ---------- Gemini call ----------
def build_prompt(url: str, categories: list[str], page_text: str) -> str:
    cats_csv = ", ".join(categories + ["unknown"])
    return f"""You are classifying a website for a directory of VENDORS/SERVICE PROVIDERS who sell to self-storage facility owners.

Vendors serve the self-storage industry with things like:
- Construction, steel buildings, doors, hardware
- Security systems, access control, surveillance cameras, gates
- HVAC, climate control, dehumidifiers
- Insurance, tenant protection programs
- Management software, websites, payment processing, marketing
- Signage, lighting, moving supplies
- Consulting, brokerage, management services
- Cleaning, pest control, maintenance

NOT vendors (reject):
- Self-storage FACILITY OPERATORS (they rent units to the public)
- Unrelated businesses (residential HVAC, retail sign makers, generic locksmiths, law firms not specializing in storage, etc.)
- News sites, blogs, forums, directories

TASK: Based on the website text below, answer in JSON.

Page URL: {url}

Page text (first {MAX_PAGE_TEXT_CHARS} chars):
---
{page_text}
---

Output JSON fields:
- relevance: "YES" (storage-industry vendor), "NO" (not storage-industry OR is a facility operator OR unrelated), or "UNCLEAR" (page has too little info to decide)
- category_slug: one of [{cats_csv}]. If relevance is NO or UNCLEAR, use "unknown".
- reason: ONE short sentence explaining your answer."""


def call_gemini(client: httpx.Client, api_key: str, prompt: str) -> dict:
    """Call Gemini with retry/backoff. Returns parsed dict or {"_error": ...}.

    Retry policy:
      - 503 (service overloaded): wait 15s, retry up to 2 times.
      - 429 (rate limited): wait 65s, retry once. If still 429 after retry,
        surface as RATE_LIMITED so caller can decide to stop (daily cap).
    """
    body = {
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {
            "responseMimeType": "application/json",
            "responseSchema": {
                "type": "object",
                "properties": {
                    "relevance": {"type": "string", "enum": ["YES", "NO", "UNCLEAR"]},
                    "category_slug": {"type": "string"},
                    "reason": {"type": "string"},
                },
                "required": ["relevance", "category_slug", "reason"],
            },
            "temperature": 0.1,
        },
    }

    max_503_retries = 2
    max_429_retries = 1
    retries_503 = 0
    retries_429 = 0

    while True:
        try:
            r = client.post(
                GEMINI_ENDPOINT,
                headers={"Content-Type": "application/json", "x-goog-api-key": api_key},
                json=body,
                timeout=GEMINI_TIMEOUT,
            )
        except httpx.HTTPError as e:
            return {"_error": f"{type(e).__name__}: {str(e)[:200]}"}

        if r.status_code == 503:
            if retries_503 < max_503_retries:
                retries_503 += 1
                print(f"    ...503 overloaded, sleeping 15s (retry {retries_503}/{max_503_retries})")
                time.sleep(15)
                continue
            return {"_error": "http 503 (overloaded, retries exhausted)", "_body": r.text[:500]}

        if r.status_code == 429:
            if retries_429 < max_429_retries:
                retries_429 += 1
                print(f"    ...429 rate-limited, sleeping 65s (retry {retries_429}/{max_429_retries})")
                time.sleep(65)
                continue
            return {"_error": "RATE_LIMITED", "_status": 429, "_body": r.text[:500]}

        if r.status_code >= 400:
            return {"_error": f"http {r.status_code}", "_body": r.text[:500]}

        try:
            data = r.json()
            text = data["candidates"][0]["content"]["parts"][0]["text"]
            return json.loads(text)
        except Exception as e:
            return {"_error": f"parse: {type(e).__name__}: {str(e)[:200]}", "_body": r.text[:500]}


# ---------- main ----------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=0,
                    help="Only process this many candidates (smoke test). 0 = all.")
    args = ap.parse_args()

    load_env_local()
    api_key = os.environ.get("GEMINI_API_KEY", "").strip()
    if not api_key:
        print("ERROR: GEMINI_API_KEY missing from .env.local", file=sys.stderr)
        return 1

    if not INPUT.exists():
        print(f"ERROR: run step 3.5 first — {INPUT} not found", file=sys.stderr)
        return 1

    print("Fetching categories from Supabase...")
    categories = fetch_categories()
    print(f"  {len(categories)} categories: {', '.join(categories)}")

    print("Fetching existing vendor domains from Supabase...")
    existing_domains = fetch_existing_vendor_domains()
    print(f"  {len(existing_domains)} existing domains")

    print(f"Reading {INPUT} ...")
    in_header, in_rows = read_xlsx(INPUT)
    idx = {h: i for i, h in enumerate(in_header)}
    if "domain" not in idx:
        print("ERROR: input missing 'domain' column", file=sys.stderr)
        return 1

    # Already-processed domains (for resume): read from all three output files
    processed: set[str] = set()
    for path in (OUT_RELEVANT, OUT_NOT_RELEVANT, OUT_REVIEW):
        _h, rows = read_xlsx(path)
        if not _h:
            continue
        d_col = _h.index("domain") if "domain" in _h else 0
        for row in rows:
            d = (row[d_col] or "").strip().lower()
            if d:
                processed.add(d)
    if processed:
        print(f"Resuming: {len(processed)} domains already processed in prior runs")

    # Build job list: skip pre-filtered + already-processed
    jobs: list[tuple[str, str]] = []  # (domain, final_url)
    prefilter_tld: list[list] = []
    prefilter_existing: list[list] = []
    for row in in_rows:
        d = (row[idx["domain"]] or "").strip().lower()
        if not d or d in processed:
            continue
        if any(d.endswith(s) for s in BLOCKED_TLD_SUFFIXES):
            prefilter_tld.append(list(row))
            continue
        if d in existing_domains:
            prefilter_existing.append(list(row))
            continue
        url = (row[idx["final_url"]] if "final_url" in idx else "") or f"https://{d}"
        jobs.append((d, url, list(row)))

    print(f"\nPre-filter (no API calls):")
    print(f"  {len(prefilter_tld)} dropped — blocked TLD")
    print(f"  {len(prefilter_existing)} dropped — already in vendors table")
    print(f"  {len(jobs)} going to Gemini")

    if args.limit:
        jobs = jobs[: args.limit]
        print(f"  (smoke test: limited to {args.limit})")

    # Output schema
    extra_cols = ["relevance", "category_guess", "reason", "api_error"]
    out_header = in_header + extra_cols

    # Reload existing output rows so we can append and checkpoint
    def load_out(path: Path) -> list[list]:
        _h, rows = read_xlsx(path)
        return rows

    rel_rows = load_out(OUT_RELEVANT)
    not_rel_rows = load_out(OUT_NOT_RELEVANT)
    rev_rows = load_out(OUT_REVIEW)

    # Add pre-filter drops to the not_relevant file (with appropriate reason)
    for row in prefilter_tld:
        not_rel_rows.append(row + ["NO", "unknown", "pre-filter: blocked TLD", ""])
    for row in prefilter_existing:
        not_rel_rows.append(row + ["NO", "unknown", "pre-filter: already in vendors table", ""])

    # Save pre-filter drops immediately so we don't re-do them if interrupted
    if prefilter_tld or prefilter_existing:
        write_xlsx(OUT_NOT_RELEVANT, out_header, not_rel_rows)

    if not jobs:
        print("\nNothing to process.")
        write_xlsx(OUT_RELEVANT, out_header, rel_rows)
        write_xlsx(OUT_NOT_RELEVANT, out_header, not_rel_rows)
        write_xlsx(OUT_REVIEW, out_header, rev_rows)
        return 0

    print(f"\nProcessing {len(jobs)} sites (pace ~{REQUEST_INTERVAL_SECONDS}s between calls)...")
    with httpx.Client(verify=False, http2=True) as client:
        for i, (domain, url, row) in enumerate(jobs, start=1):
            tick = time.time()

            page_text, fetch_err = fetch_page_text(client, domain, url)
            if fetch_err:
                rev_rows.append(row + ["UNCLEAR", "unknown", f"fetch failed: {fetch_err}", fetch_err])
                print(f"[{i}/{len(jobs)}] {domain} -> UNCLEAR (fetch failed: {fetch_err})")
            else:
                prompt = build_prompt(url, categories, page_text)
                result = call_gemini(client, api_key, prompt)

                if result.get("_error") == "RATE_LIMITED":
                    # On paid tier this should be rare. Log as UNCLEAR so we
                    # don't bail on the whole run; resume will retry later.
                    rev_rows.append(row + ["UNCLEAR", "unknown", "rate limited (retry later)", "429"])
                    print(f"[{i}/{len(jobs)}] {domain} -> UNCLEAR (rate limited)")
                elif "_error" in result:
                    err = result["_error"]
                    rev_rows.append(row + ["UNCLEAR", "unknown", f"api error: {err}", err])
                    print(f"[{i}/{len(jobs)}] {domain} -> UNCLEAR (api error: {err})")
                else:
                    rel = str(result.get("relevance", "UNCLEAR")).upper()
                    cat = str(result.get("category_slug", "unknown"))
                    reason = str(result.get("reason", ""))
                    out_row = row + [rel, cat, reason, ""]
                    if rel == "YES":
                        rel_rows.append(out_row)
                        bucket = "relevant"
                    elif rel == "NO":
                        not_rel_rows.append(out_row)
                        bucket = "not relevant"
                    else:
                        rev_rows.append(out_row)
                        bucket = "needs review"
                    print(f"[{i}/{len(jobs)}] {domain} -> {rel} [{cat}] ({bucket})")

            # Checkpoint every 20 rows and at the end
            if i % 20 == 0 or i == len(jobs):
                write_xlsx(OUT_RELEVANT, out_header, rel_rows)
                write_xlsx(OUT_NOT_RELEVANT, out_header, not_rel_rows)
                write_xlsx(OUT_REVIEW, out_header, rev_rows)

            # Pace requests to stay under 15 RPM
            elapsed = time.time() - tick
            if elapsed < REQUEST_INTERVAL_SECONDS:
                time.sleep(REQUEST_INTERVAL_SECONDS - elapsed)

    # Final write (belt + suspenders)
    write_xlsx(OUT_RELEVANT, out_header, rel_rows)
    write_xlsx(OUT_NOT_RELEVANT, out_header, not_rel_rows)
    write_xlsx(OUT_REVIEW, out_header, rev_rows)

    print(f"\nDone.")
    print(f"  relevant:     {len(rel_rows)}")
    print(f"  not_relevant: {len(not_rel_rows)}")
    print(f"  needs_review: {len(rev_rows)}")
    print(f"\nWrote {OUT_RELEVANT}")
    print(f"Wrote {OUT_NOT_RELEVANT}")
    print(f"Wrote {OUT_REVIEW}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
