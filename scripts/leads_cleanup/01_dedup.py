"""
STEP 1 — Deduplicate leads.xlsx by website domain.

Input:  templates/leads.xlsx (26k+ raw rows from Serper)
Output: templates/01_leads_deduped.xlsx (~5k unique domains)

For each unique domain we keep ONE "best" row, picked by:
  1. Source quality: knowledgeGraph > places > organic
  2. Phone present > phone missing
  3. Address present > address missing
  4. Lowest rank within the source

We also add two new columns so no information is lost:
  - seen_in_queries    : comma-separated list of every query that surfaced this domain
  - query_count        : how many queries surfaced it (high = generic/aggregator, low = specific)

Run:
    python scripts/leads_cleanup/01_dedup.py
"""

import sys
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
# Path to the raw leads file produced by serper_leads.py.
# Adjust if you move the leads file to a different location.
INPUT = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates\leads.xlsx"
)
OUTPUT = INPUT.parent / "01_leads_deduped.xlsx"

# Source-quality ranking (lower = better)
SOURCE_RANK = {"knowledgeGraph": 0, "places": 1, "organic": 2}

HEADERS_OUT = [
    "domain", "name_from_serper", "website", "phone", "address",
    "best_source", "best_rank", "query_count", "seen_in_queries",
    "best_snippet",
]


def normalize_domain(url: str) -> str:
    if not url:
        return ""
    try:
        host = urlparse(url).netloc.lower()
    except Exception:
        return ""
    return host.removeprefix("www.")


def row_score(source: str, rank, phone, address) -> tuple:
    """Lower tuple = better row. Used to break ties and pick the best row per domain."""
    return (
        SOURCE_RANK.get(source, 9),        # source quality
        0 if phone else 1,                  # phone present wins
        0 if address else 1,                # address present wins
        int(rank) if isinstance(rank, (int, float)) else 99,
    )


def main() -> int:
    if not INPUT.exists():
        print(f"ERROR: input file not found: {INPUT}", file=sys.stderr)
        return 1

    print(f"Reading {INPUT} ...")
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb.active

    # best[domain] = (score_tuple, record_dict, queries_set)
    best: dict[str, tuple] = {}
    total = 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 8:
            continue
        q, rank, name, url, snippet, phone, addr, src = row[:8]
        if not url:
            continue
        domain = normalize_domain(url)
        if not domain:
            continue
        total += 1

        score = row_score(src or "", rank, phone, addr)
        if domain not in best:
            best[domain] = (
                score,
                {
                    "domain": domain,
                    "name_from_serper": (name or "").strip(),
                    "website": f"https://{domain}",   # canonical homepage URL
                    "phone": phone or "",
                    "address": addr or "",
                    "best_source": src or "",
                    "best_rank": rank if rank is not None else "",
                    "best_snippet": (snippet or "").strip(),
                },
                {q or ""},
            )
        else:
            prev_score, prev_rec, queries = best[domain]
            if score < prev_score:
                best[domain] = (
                    score,
                    {
                        "domain": domain,
                        "name_from_serper": (name or "").strip(),
                        "website": f"https://{domain}",
                        "phone": phone or "",
                        "address": addr or "",
                        "best_source": src or "",
                        "best_rank": rank if rank is not None else "",
                        "best_snippet": (snippet or "").strip(),
                    },
                    queries,
                )
            queries.add(q or "")

    print(f"Read {total} rows -> {len(best)} unique domains")

    # Write output
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Deduped"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(HEADERS_OUT, start=1):
        cell = out_ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    widths = {
        "domain": 28, "name_from_serper": 36, "website": 34, "phone": 16,
        "address": 36, "best_source": 14, "best_rank": 8, "query_count": 10,
        "seen_in_queries": 60, "best_snippet": 60,
    }
    for c, h in enumerate(HEADERS_OUT, start=1):
        out_ws.column_dimensions[get_column_letter(c)].width = widths.get(h, 20)

    # Sort: fewest queries first (more specific = more likely a real niche vendor),
    # then alphabetical. You can re-sort in Excel later.
    sorted_items = sorted(
        best.items(),
        key=lambda kv: (len(kv[1][2]), kv[0]),
    )

    r = 2
    for domain, (_score, rec, queries) in sorted_items:
        rec["query_count"] = len(queries)
        rec["seen_in_queries"] = " | ".join(sorted(q for q in queries if q))
        for c, h in enumerate(HEADERS_OUT, start=1):
            out_ws.cell(row=r, column=c, value=rec.get(h, ""))
        r += 1

    out_ws.freeze_panes = "A2"
    out_wb.save(OUTPUT)
    print(f"Wrote {OUTPUT}")
    print(f"  {len(best)} unique domains")
    print(f"  {sum(1 for _, (_, _, qs) in best.items() if len(qs) == 1)} "
          f"appear in only one query")
    print(f"  {sum(1 for _, (_, _, qs) in best.items() if len(qs) >= 10)} "
          f"appear in 10+ queries (likely aggregators)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
