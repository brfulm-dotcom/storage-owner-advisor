"""
STEP 3b — Retry 403-blocked domains with browser-realistic headers.

Many vendor sites return HTTP 403 to our vanilla reachability check because
they run bot protection (Cloudflare, Akamai, etc.) that inspects request
headers. Adding full Chrome-style headers usually gets us through.

Input:   templates/03_leads_live.xlsx
         templates/03_leads_dead.xlsx

Output:  templates/03b_leads_live.xlsx   (original live + recovered 403s)
         templates/03b_leads_dead.xlsx   (still-dead, including un-recovered 403s)

Subsequent steps should read the "03b_" files, not "03_".

Run:
    python scripts/leads_cleanup/03b_retry_403.py
"""

import asyncio
import random
import sys
from pathlib import Path
from urllib.parse import urlparse

import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
BASE_DIR = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates"
)
IN_LIVE = BASE_DIR / "03_leads_live.xlsx"
IN_DEAD = BASE_DIR / "03_leads_dead.xlsx"
OUT_LIVE = BASE_DIR / "03b_leads_live.xlsx"
OUT_DEAD = BASE_DIR / "03b_leads_dead.xlsx"

CONCURRENCY = 10            # lower than step 3 — be polite with retries
TIMEOUT_SECONDS = 20

# Pool of realistic desktop Chrome User-Agents (rotated per-request)
USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]


def browser_headers(host: str) -> dict:
    """Return a fresh set of Chrome-like request headers."""
    ua = random.choice(USER_AGENTS)
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
                  "image/avif,image/webp,image/apng,*/*;q=0.8,"
                  "application/signed-exchange;v=b3;q=0.7",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Cache-Control": "no-cache",
        "Pragma": "no-cache",
        "Sec-Ch-Ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Upgrade-Insecure-Requests": "1",
        "Referer": f"https://www.google.com/search?q={host}",
    }


def normalize(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


async def retry_one(client: httpx.AsyncClient, domain: str) -> dict:
    """Retry a 403-blocked domain with browser-realistic headers."""
    for scheme in ("https", "http"):
        try:
            r = await client.get(
                f"{scheme}://{domain}",
                follow_redirects=True,
                headers=browser_headers(domain),
                timeout=TIMEOUT_SECONDS,
            )
            return {
                "status_code": r.status_code,
                "final_url": str(r.url),
                "final_domain": normalize(str(r.url)),
                "error": "",
            }
        except httpx.HTTPError as e:
            last_err = f"{type(e).__name__}: {str(e)[:120]}"
        except Exception as e:
            last_err = f"{type(e).__name__}: {str(e)[:120]}"
    return {"status_code": 0, "final_url": "", "final_domain": "", "error": last_err}


async def run(domains: list[str]) -> dict[str, dict]:
    limits = httpx.Limits(max_keepalive_connections=CONCURRENCY, max_connections=CONCURRENCY)
    sem = asyncio.Semaphore(CONCURRENCY)
    results: dict[str, dict] = {}

    async with httpx.AsyncClient(limits=limits, verify=False, http2=True) as client:
        async def worker(d: str):
            async with sem:
                # small jitter so we don't hammer one bot-protection provider at once
                await asyncio.sleep(random.uniform(0.2, 0.6))
                results[d] = await retry_one(client, d)

        tasks = [asyncio.create_task(worker(d)) for d in domains]
        total = len(tasks)
        done = 0
        for t in asyncio.as_completed(tasks):
            await t
            done += 1
            if done % 25 == 0 or done == total:
                print(f"  {done}/{total} retried")

    return results


def read_xlsx(path: Path) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
    return header, rows


def write_xlsx(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(str(h)) + 4, 40))
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    if not IN_LIVE.exists() or not IN_DEAD.exists():
        print(f"ERROR: run step 3 first — missing {IN_LIVE} or {IN_DEAD}", file=sys.stderr)
        return 1

    print(f"Reading {IN_LIVE} ...")
    live_header, live_rows = read_xlsx(IN_LIVE)

    print(f"Reading {IN_DEAD} ...")
    dead_header, dead_rows = read_xlsx(IN_DEAD)

    # Locate relevant columns in dead file
    d_idx = {h: i for i, h in enumerate(dead_header)}
    if "status_code" not in d_idx or "domain" not in d_idx:
        print("ERROR: dead file missing 'status_code' or 'domain' columns", file=sys.stderr)
        return 1

    # Split dead rows into "retry these" (403) vs "leave alone"
    retry_rows: list[list] = []
    other_dead_rows: list[list] = []
    for row in dead_rows:
        sc = row[d_idx["status_code"]]
        if sc == 403 or sc == "403":
            retry_rows.append(row)
        else:
            other_dead_rows.append(row)

    print(f"  {len(retry_rows)} rows to retry (status 403)")
    print(f"  {len(other_dead_rows)} rows left as-is")

    if not retry_rows:
        print("Nothing to retry. Copying files through unchanged.")
        write_xlsx(OUT_LIVE, live_header, live_rows)
        write_xlsx(OUT_DEAD, dead_header, dead_rows)
        print(f"Wrote {OUT_LIVE}")
        print(f"Wrote {OUT_DEAD}")
        return 0

    domains_to_retry = [row[d_idx["domain"]].strip().lower() for row in retry_rows]

    print(f"Retrying with browser-realistic headers "
          f"(concurrency={CONCURRENCY}, timeout={TIMEOUT_SECONDS}s)...")
    results = asyncio.run(run(domains_to_retry))

    # Build maps of which final_domains are already live (avoid dupes from redirects)
    live_idx = {h: i for i, h in enumerate(live_header)}
    seen_final: set[str] = set()
    if "final_domain" in live_idx:
        for r in live_rows:
            fd = (r[live_idx["final_domain"]] or "").strip().lower()
            if fd:
                seen_final.add(fd)

    # Partition retry rows into "recovered" and "still dead"
    recovered_rows: list[list] = []
    still_dead_rows: list[list] = []
    recovered = 0
    for row, d in zip(retry_rows, domains_to_retry):
        res = results.get(d, {"status_code": 0, "final_url": "", "final_domain": "", "error": "no result"})
        status = res["status_code"]
        final = res["final_domain"] or d
        new_row = list(row)
        # Update the trailing columns that step 3 added
        for col in ("status_code", "final_url", "final_domain", "error", "reason"):
            if col in d_idx:
                if col == "status_code":
                    new_row[d_idx[col]] = status
                elif col == "final_url":
                    new_row[d_idx[col]] = res["final_url"]
                elif col == "final_domain":
                    new_row[d_idx[col]] = final
                elif col == "error":
                    new_row[d_idx[col]] = ""
                elif col == "reason":
                    new_row[d_idx[col]] = ""

        if status and 200 <= status < 400:
            if final in seen_final:
                # Redirect collides with something already live
                if "reason" in d_idx:
                    new_row[d_idx["reason"]] = "redirected to duplicate (recovered)"
                still_dead_rows.append(new_row)
            else:
                seen_final.add(final)
                recovered_rows.append(new_row)
                recovered += 1
        else:
            # Still blocked / errored — mark with final reason
            if "reason" in d_idx:
                if status >= 400:
                    new_row[d_idx["reason"]] = f"still http {status} after retry"
                else:
                    new_row[d_idx["reason"]] = f"unreachable after retry: {res['error']}"
            still_dead_rows.append(new_row)

    print(f"  recovered: {recovered} (now live)")
    print(f"  still dead: {len(still_dead_rows)}")

    # Merge recovered rows into live set. The schemas differ — live rows don't have
    # a 'reason' column. Slice recovered rows to the live header length.
    live_extended = list(live_rows)
    trim_len = len(live_header)
    for r in recovered_rows:
        live_extended.append(r[:trim_len])

    # Final dead = the already-non-403 rows + anything still dead after retry
    final_dead = other_dead_rows + still_dead_rows

    write_xlsx(OUT_LIVE, live_header, live_extended)
    write_xlsx(OUT_DEAD, dead_header, final_dead)

    print(f"\nWrote {OUT_LIVE}  ({len(live_extended)} rows)")
    print(f"Wrote {OUT_DEAD}  ({len(final_dead)} rows)")
    print(f"\nTotal recovery rate: {recovered}/{len(retry_rows)} "
          f"({100 * recovered / max(1, len(retry_rows)):.1f}%)")
    return 0


if __name__ == "__main__":
    sys.exit(main())
