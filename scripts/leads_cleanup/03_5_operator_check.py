"""
STEP 3.5 — Heuristic operator detection.

Fetches each live homepage and scores it for "is this a self-storage facility
operator?" using keyword/pattern heuristics. No LLM, no API key needed.

We want to KEEP vendors/service providers and REJECT operators.

Signals (each unique signal = 1 point):
  - Rental CTAs: "rent a unit", "reserve your unit", "rent online", "move-in today"
  - Unit size pricing: "5x5", "10x10", "10x20" in pricing/rental context
  - Facility finder: "find a location", "find a facility", "our locations"
  - Facility amenities: "drive-up access", "climate controlled units", "gate hours"
  - Zip-code search box for finding a facility
  - Phrase "storage units" as primary navigation / H1
  - "first month free" / "move-in special" — rental promo language

Score buckets:
  0-1  -> vendor_candidates
  2+   -> likely_operators
  error / empty page -> unclear

Input:   templates/03b_leads_live.xlsx
Outputs: templates/03_5_vendor_candidates.xlsx
         templates/03_5_likely_operators.xlsx
         templates/03_5_unclear.xlsx

Run:
    python scripts/leads_cleanup/03_5_operator_check.py
"""

import asyncio
import random
import re
import sys
from pathlib import Path
from urllib.parse import urlparse

import httpx
import openpyxl
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
BASE_DIR = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates"
)
INPUT = BASE_DIR / "03b_leads_live.xlsx"
OUT_VENDORS = BASE_DIR / "03_5_vendor_candidates.xlsx"
OUT_OPERATORS = BASE_DIR / "03_5_likely_operators.xlsx"
OUT_UNCLEAR = BASE_DIR / "03_5_unclear.xlsx"

CONCURRENCY = 15
TIMEOUT_SECONDS = 20
OPERATOR_SCORE_THRESHOLD = 2

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36",
]

# ---------- SIGNAL PATTERNS ----------
# Each entry: (signal_name, compiled regex). Text is lowercased before matching.
SIGNAL_PATTERNS: list[tuple[str, re.Pattern]] = [
    ("rent_cta", re.compile(
        r"\b(rent a unit|rent online|reserve (your|a) unit|reserve online|"
        r"rent this unit|rent now|book online|rent storage)\b")),
    ("move_in_promo", re.compile(
        r"\b(move[- ]in (today|special|now)|first month (free|\$1)|"
        r"move in now|online move[- ]in)\b")),
    ("find_location", re.compile(
        r"\b(find a (location|facility|store|unit)|our locations|"
        r"storage (near|locations in)|nearest (facility|location)|"
        r"view (our )?locations)\b")),
    ("unit_sizes_pricing", re.compile(
        r"\b(5\s*['x×]\s*5|5\s*['x×]\s*10|10\s*['x×]\s*10|10\s*['x×]\s*15|"
        r"10\s*['x×]\s*20|10\s*['x×]\s*30)\b")),
    ("drive_up_units", re.compile(
        r"\b(drive[- ]up (access|units?|storage)|ground[- ]floor (units?|access))\b")),
    ("climate_units", re.compile(
        r"\b(climate[- ]controlled (units?|storage)|temperature[- ]controlled (units?|storage))\b")),
    ("gate_hours", re.compile(
        r"\b(gate (code|hours|access hours)|access (hours|code)|"
        r"24[- ]hour access)\b")),
    ("storage_units_nav", re.compile(
        r"\b(storage units?|self[- ]storage)\b.{0,40}\b(rental|rent|available|sizes?|prices?)\b")),
    ("zip_search", re.compile(
        r"\b(enter (your )?zip|find by zip|search by (zip|location)|zip code)\b")),
    ("hours_of_operation", re.compile(
        r"\b(office hours|hours of operation)\b")),
]


def fetch_headers(domain: str) -> dict:
    ua = random.choice(USER_AGENTS)
    return {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Upgrade-Insecure-Requests": "1",
        "Sec-Ch-Ua": '"Google Chrome";v="131", "Chromium";v="131", "Not_A Brand";v="24"',
        "Sec-Ch-Ua-Mobile": "?0",
        "Sec-Ch-Ua-Platform": '"Windows"',
        "Sec-Fetch-Dest": "document",
        "Sec-Fetch-Mode": "navigate",
        "Sec-Fetch-Site": "none",
        "Sec-Fetch-User": "?1",
        "Referer": f"https://www.google.com/search?q={domain}",
    }


def score_html(html: str) -> tuple[int, list[str]]:
    """Return (score, list of matched signal names)."""
    if not html:
        return 0, []
    try:
        soup = BeautifulSoup(html, "html.parser")
        # Remove noise: scripts, styles, hidden elements
        for tag in soup(["script", "style", "noscript"]):
            tag.decompose()
        text = soup.get_text(" ", strip=True).lower()
    except Exception:
        text = html.lower()

    matched: list[str] = []
    for name, pat in SIGNAL_PATTERNS:
        if pat.search(text):
            matched.append(name)
    return len(matched), matched


async def fetch_one(client: httpx.AsyncClient, url: str, fallback_domain: str) -> dict:
    """Fetch a URL. Returns dict with html, final_url, status, error."""
    target = url or f"https://{fallback_domain}"
    try:
        r = await client.get(
            target,
            follow_redirects=True,
            headers=fetch_headers(fallback_domain),
            timeout=TIMEOUT_SECONDS,
        )
        return {
            "status": r.status_code,
            "final_url": str(r.url),
            "html": r.text if 200 <= r.status_code < 400 else "",
            "error": "" if 200 <= r.status_code < 400 else f"http {r.status_code}",
        }
    except httpx.HTTPError as e:
        return {"status": 0, "final_url": "", "html": "", "error": f"{type(e).__name__}: {str(e)[:120]}"}
    except Exception as e:
        return {"status": 0, "final_url": "", "html": "", "error": f"{type(e).__name__}: {str(e)[:120]}"}


async def run(jobs: list[tuple[str, str]]) -> dict[str, dict]:
    """jobs is list of (domain, url_to_fetch). Returns {domain: result}."""
    limits = httpx.Limits(max_keepalive_connections=CONCURRENCY, max_connections=CONCURRENCY)
    sem = asyncio.Semaphore(CONCURRENCY)
    results: dict[str, dict] = {}

    async with httpx.AsyncClient(limits=limits, verify=False, http2=True) as client:
        async def worker(domain: str, url: str):
            async with sem:
                await asyncio.sleep(random.uniform(0.1, 0.4))
                res = await fetch_one(client, url, domain)
                score, signals = score_html(res["html"])
                results[domain] = {
                    "status": res["status"],
                    "final_url": res["final_url"],
                    "error": res["error"],
                    "score": score,
                    "signals": signals,
                }

        tasks = [asyncio.create_task(worker(d, u)) for d, u in jobs]
        total = len(tasks)
        done = 0
        for t in asyncio.as_completed(tasks):
            await t
            done += 1
            if done % 100 == 0 or done == total:
                print(f"  {done}/{total} fetched+scored")

    return results


def read_xlsx(path: Path) -> tuple[list[str], list[list]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [list(r) for r in ws.iter_rows(min_row=2, values_only=True) if r and any(r)]
    return header, rows


def write_xlsx(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(str(h)) + 4, 40))
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> int:
    if not INPUT.exists():
        print(f"ERROR: run step 3b first — {INPUT} not found", file=sys.stderr)
        return 1

    print(f"Reading {INPUT} ...")
    header, rows = read_xlsx(INPUT)
    idx = {h: i for i, h in enumerate(header)}

    if "domain" not in idx:
        print("ERROR: input missing 'domain' column", file=sys.stderr)
        return 1

    # Prefer final_url from step 3; fall back to https://domain
    jobs: list[tuple[str, str]] = []
    for row in rows:
        d = (row[idx["domain"]] or "").strip().lower()
        if not d:
            continue
        url = (row[idx["final_url"]] if "final_url" in idx else "") or f"https://{d}"
        jobs.append((d, url))

    print(f"Fetching + scoring {len(jobs)} live sites "
          f"(concurrency={CONCURRENCY}, timeout={TIMEOUT_SECONDS}s)...")
    results = asyncio.run(run(jobs))

    # Partition rows into three buckets
    extra_cols = ["operator_score", "matched_signals", "fetch_error"]
    out_header = header + extra_cols

    vendor_rows: list[list] = []
    operator_rows: list[list] = []
    unclear_rows: list[list] = []

    for row in rows:
        d = (row[idx["domain"]] or "").strip().lower()
        if not d:
            continue
        res = results.get(d, {"score": 0, "signals": [], "error": "no result"})
        score = res["score"]
        signals_csv = ", ".join(res["signals"])
        err = res["error"]
        new_row = list(row) + [score, signals_csv, err]

        if err and not res["signals"]:
            unclear_rows.append(new_row)
        elif score >= OPERATOR_SCORE_THRESHOLD:
            operator_rows.append(new_row)
        else:
            vendor_rows.append(new_row)

    print(f"\nResults:")
    print(f"  vendor_candidates: {len(vendor_rows)}")
    print(f"  likely_operators:  {len(operator_rows)}")
    print(f"  unclear:           {len(unclear_rows)}")

    write_xlsx(OUT_VENDORS, out_header, vendor_rows)
    write_xlsx(OUT_OPERATORS, out_header, operator_rows)
    write_xlsx(OUT_UNCLEAR, out_header, unclear_rows)
    print(f"\nWrote {OUT_VENDORS}")
    print(f"Wrote {OUT_OPERATORS}")
    print(f"Wrote {OUT_UNCLEAR}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
