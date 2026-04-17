"""
STEP 2 — Block-list filter.

Removes:
  1. Domains of vendors we already have in Supabase (no point re-importing).
  2. Directory / aggregator / review / social-media sites that are not real vendors.
  3. Government / military / education TLDs — .gov, .mil, .edu plus foreign
     equivalents (.gov.uk, .ac.uk, .edu.au, .gc.ca, etc.). These aren't vendors.

Note: the old "generic platform" frequency check (reject domains appearing in 50+
queries) was removed. Many of those are legitimate national vendors. Known big
platforms we want blocked are listed explicitly in STATIC_BLOCK instead.

Input:  templates/01_leads_deduped.xlsx
Output: templates/02_leads_filtered.xlsx  (survivors — go to reachability step)
        templates/02_leads_rejected.xlsx  (what got cut, with reason — for your audit)

Reads Supabase credentials from:
  storage-owner-advisor/.env.local   (NEXT_PUBLIC_SUPABASE_URL + SUPABASE_SERVICE_ROLE_KEY)

Run:
    python scripts/leads_cleanup/02_blocklist.py
"""

import os
import sys
from pathlib import Path
from urllib.parse import urlparse

import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
INPUT = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates\01_leads_deduped.xlsx"
)
OUT_KEEP = INPUT.parent / "02_leads_filtered.xlsx"
OUT_REJ = INPUT.parent / "02_leads_rejected.xlsx"

ENV_PATH = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor\.env.local"
)

# Hard-coded block list. Edit as you notice more junk in the rejected file.
STATIC_BLOCK = {
    # Social media / big tech
    "facebook.com", "linkedin.com", "twitter.com", "x.com", "instagram.com",
    "youtube.com", "tiktok.com", "pinterest.com", "reddit.com",
    "google.com", "bing.com", "apple.com", "amazon.com", "wikipedia.org",

    # Review / directory sites
    "yelp.com", "bbb.org", "yellowpages.com", "mapquest.com", "tripadvisor.com",
    "angi.com", "trustpilot.com", "g2.com", "capterra.com", "softwareadvice.com",
    "glassdoor.com", "indeed.com", "thumbtack.com",

    # Our own site
    "storageowneradvisor.com",

    # Large generic platforms / competitors — these are not new leads for you.
    "storable.com", "sitelink.com", "rentcafe.com", "storagepug.com",
    "insideselfstorage.com", "nsastorage.com", "modernstoragemedia.com",

    # Large operators (user decision: reject operators from this import)
    "extraspace.com", "publicstorage.com", "uhaul.com", "cubesmart.com",
    "life-storage.com", "lifestorage.com", "simplyss.com", "storagemart.com",
    "stopandstor.com", "stop-and-stor.com",
}

# Government / military / education TLD suffixes — not vendors.
# A domain ending with any of these is rejected. Leading dot is required.
BLOCKED_TLD_SUFFIXES = (
    # US government / military / education (primary)
    ".gov", ".mil", ".edu",

    # Foreign government variants
    ".gov.uk", ".gov.au", ".gov.ca", ".gov.nz", ".gov.in", ".gov.za",
    ".gov.sg", ".gov.ie", ".gov.hk", ".gov.jp", ".gov.tw", ".gov.il",
    ".gov.br", ".gov.mx", ".gouv.fr",
    ".gc.ca",  # Canadian federal

    # Foreign military variants
    ".mil.uk", ".mil.au", ".mil.nz", ".mil.za", ".mil.br",

    # Foreign education / academic variants
    ".edu.au", ".edu.cn", ".edu.hk", ".edu.sg", ".edu.in", ".edu.tw",
    ".edu.mx", ".edu.br", ".edu.my",
    ".ac.uk", ".ac.nz", ".ac.in", ".ac.jp", ".ac.za", ".ac.kr", ".ac.ir", ".ac.il",
)


def has_blocked_tld(domain: str) -> str | None:
    """Return the matched TLD suffix if `domain` ends with one, else None."""
    for suffix in BLOCKED_TLD_SUFFIXES:
        if domain.endswith(suffix):
            return suffix
    return None


def load_env_local():
    if not ENV_PATH.exists():
        return
    for line in ENV_PATH.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))


def load_existing_vendor_domains() -> set[str]:
    """Fetch all vendor.website values from Supabase and return normalized domains."""
    url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL", "").rstrip("/")
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY", "")
    if not url or not key:
        print("WARN: Supabase env vars missing; skipping 'already-in-database' check.",
              file=sys.stderr)
        return set()

    print("Fetching existing vendor domains from Supabase...")
    # Page through in chunks of 1000 to be safe
    all_websites: list[str] = []
    offset = 0
    page = 1000
    with httpx.Client(timeout=30) as client:
        while True:
            r = client.get(
                f"{url}/rest/v1/vendors",
                params={"select": "website", "limit": str(page), "offset": str(offset)},
                headers={"apikey": key, "Authorization": f"Bearer {key}"},
            )
            r.raise_for_status()
            rows = r.json()
            if not rows:
                break
            all_websites.extend((row.get("website") or "") for row in rows)
            if len(rows) < page:
                break
            offset += page

    domains = set()
    for w in all_websites:
        if not w:
            continue
        try:
            h = urlparse(w).netloc.lower().removeprefix("www.")
        except Exception:
            continue
        if h:
            domains.add(h)
    print(f"  Loaded {len(domains)} existing vendor domains")
    return domains


def main() -> int:
    if not INPUT.exists():
        print(f"ERROR: run step 1 first — {INPUT} not found", file=sys.stderr)
        return 1

    load_env_local()
    existing = load_existing_vendor_domains()

    print(f"Reading {INPUT} ...")
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    keep_rows: list[list] = []
    reject_rows: list[list] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        domain = (row[idx["domain"]] or "").strip().lower()
        if not domain:
            continue

        reason = None
        if domain in existing:
            reason = "already in database"
        elif domain in STATIC_BLOCK:
            reason = "static block list"
        elif any(domain.endswith("." + b) for b in STATIC_BLOCK):
            reason = "subdomain of blocked site"
        elif (tld := has_blocked_tld(domain)):
            reason = f"blocked TLD ({tld})"

        if reason:
            reject_rows.append(list(row) + [reason])
        else:
            keep_rows.append(list(row))

    print(f"  kept:     {len(keep_rows)}")
    print(f"  rejected: {len(reject_rows)}")

    write_xlsx(OUT_KEEP, header, keep_rows)
    write_xlsx(OUT_REJ, header + ["rejection_reason"], reject_rows)
    print(f"Wrote {OUT_KEEP}")
    print(f"Wrote {OUT_REJ}")
    return 0


def write_xlsx(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(str(h)) + 4, 40))
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
