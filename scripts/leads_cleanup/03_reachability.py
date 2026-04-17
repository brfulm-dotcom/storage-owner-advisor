"""
STEP 3 — Reachability check.

For each surviving domain, make a quick HTTP request to confirm the site is live
and follow redirects to see the canonical URL. Dead sites are skipped before we
spend money on LLM enrichment.

Input:  templates/02_leads_filtered.xlsx
Output: templates/03_leads_live.xlsx   (status 200, ready for enrichment)
        templates/03_leads_dead.xlsx   (dead / redirected away / errored)

Also re-dedupes after redirects: if three different leads all redirect to the same
parent domain, we keep only one live row and move the others to dead with a
"redirected to duplicate" reason.

Run:
    python scripts/leads_cleanup/03_reachability.py
"""

import asyncio
import sys
from pathlib import Path
from urllib.parse import urlparse

import httpx
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ---------- CONFIG ----------
INPUT = Path(
    r"C:\Projects\StorageOwnerAdvisor-project\storage-owner-advisor"
    r"\.claude\worktrees\friendly-sinoussi\templates\02_leads_filtered.xlsx"
)
OUT_LIVE = INPUT.parent / "03_leads_live.xlsx"
OUT_DEAD = INPUT.parent / "03_leads_dead.xlsx"

CONCURRENCY = 20            # how many requests in flight at once
TIMEOUT_SECONDS = 15
USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
)


def normalize(url: str) -> str:
    try:
        return urlparse(url).netloc.lower().removeprefix("www.")
    except Exception:
        return ""


async def check_one(client: httpx.AsyncClient, domain: str) -> dict:
    """Try GET on https then fallback to http. Return status + final domain."""
    for scheme in ("https", "http"):
        try:
            r = await client.get(
                f"{scheme}://{domain}",
                follow_redirects=True,
                headers={"User-Agent": USER_AGENT},
                timeout=TIMEOUT_SECONDS,
            )
            final_domain = normalize(str(r.url))
            return {
                "status_code": r.status_code,
                "final_url": str(r.url),
                "final_domain": final_domain,
                "error": "",
            }
        except httpx.HTTPError as e:
            last_err = f"{type(e).__name__}: {str(e)[:120]}"
        except Exception as e:
            last_err = f"{type(e).__name__}: {str(e)[:120]}"
    return {"status_code": 0, "final_url": "", "final_domain": "", "error": last_err}


async def run(domains: list[str]) -> dict[str, dict]:
    limits = httpx.Limits(max_keepalive_connections=CONCURRENCY, max_connections=CONCURRENCY)
    results: dict[str, dict] = {}
    sem = asyncio.Semaphore(CONCURRENCY)

    async with httpx.AsyncClient(limits=limits, verify=False) as client:
        async def worker(d: str):
            async with sem:
                r = await check_one(client, d)
                results[d] = r

        total = len(domains)
        tasks = [asyncio.create_task(worker(d)) for d in domains]
        done = 0
        for t in asyncio.as_completed(tasks):
            await t
            done += 1
            if done % 100 == 0 or done == total:
                print(f"  {done}/{total} checked")

    return results


def main() -> int:
    if not INPUT.exists():
        print(f"ERROR: run step 2 first — {INPUT} not found", file=sys.stderr)
        return 1

    print(f"Reading {INPUT} ...")
    wb = openpyxl.load_workbook(INPUT, read_only=True, data_only=True)
    ws = wb.active
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {h: i for i, h in enumerate(header)}

    rows = []
    domains = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not row[idx["domain"]]:
            continue
        rows.append(list(row))
        domains.append(row[idx["domain"]].strip().lower())

    print(f"Checking reachability of {len(domains)} domains "
          f"(concurrency={CONCURRENCY}, timeout={TIMEOUT_SECONDS}s)...")
    results = asyncio.run(run(domains))

    # Classify + re-dedupe on final domain
    live_rows: list[list] = []
    dead_rows: list[list] = []
    seen_final_domains: set[str] = set()

    extra_cols = ["status_code", "final_url", "final_domain", "error", "reason"]

    for row, d in zip(rows, domains):
        res = results.get(d, {"status_code": 0, "final_url": "", "final_domain": "", "error": "no result"})
        status = res["status_code"]
        final = res["final_domain"] or d

        if status and 200 <= status < 400:
            if final in seen_final_domains:
                dead_rows.append(row + [status, res["final_url"], final, "", "redirected to duplicate"])
            else:
                seen_final_domains.add(final)
                live_rows.append(row + [status, res["final_url"], final, "", ""])
        elif status >= 400:
            dead_rows.append(row + [status, res["final_url"], final, "", f"http {status}"])
        else:
            dead_rows.append(row + [0, "", "", res["error"], "unreachable"])

    print(f"  live: {len(live_rows)}")
    print(f"  dead: {len(dead_rows)}")

    write_xlsx(OUT_LIVE, header + extra_cols, live_rows)
    write_xlsx(OUT_DEAD, header + extra_cols, dead_rows)
    print(f"Wrote {OUT_LIVE}")
    print(f"Wrote {OUT_DEAD}")
    return 0


def write_xlsx(path: Path, header: list[str], rows: list[list]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Rows"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for c, h in enumerate(header, start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(c)].width = max(14, min(len(str(h)) + 4, 40))
    for r, row in enumerate(rows, start=2):
        for c, v in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=v)
    ws.freeze_panes = "A2"
    wb.save(path)


if __name__ == "__main__":
    sys.exit(main())
